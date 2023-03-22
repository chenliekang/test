import csv
import openpyxl
import xlrd
import traceback


# 为智家按要求解析csv文件
def get_repeat_tiame(path, header, column_index, phone_range=[False, 0]):
    # path：csv文档存储路径
    # header：csv文档的标题
    # column_index：统计哪一列
    # phone_range：截取号码字段
    result = {}     # 解析后返回的结果
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if not phone_range[0]:
                phone_range[1] = len(row[column_index])
            if header in row[column_index]:
                # 是标题的话不存入结果resulr
                pass
            elif row[column_index][:phone_range[1]] in result.keys():
                # 若该字段result的键里已有，则数量+1
                result[row[column_index][:phone_range[1]]] += 1
            else:
                # 若该字段result的键里没有，则新建该字段并数量为1
                result[row[column_index][:phone_range[1]]] = 1
    return result


# 为智家按要求解析csv文件（关键字：号码+时间）
def get_repeat_tiame2(path, header):
    # path：csv文档存储路径
    # header：csv文档的标题
    # column_index：统计哪一列
    # phone_range：截取号码字段
    result = {}     # 解析后返回的结果
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if header in row[0]:
                # 是标题的话不存入结果resulr
                pass
            elif row[0]+'_'+row[2] in result.keys():
                # 若该字段result的键里已有，则数量+1
                result[row[0]+'_'+row[2]] += 1
            else:
                # 若该字段result的键里没有，则新建该字段并数量为1
                result[row[0]+'_'+row[2]] = 1
    return result


# 为智家把解析后的结果存入表格中
def set_result(result, lower_limit, path, titles):
    # result：解析结果
    # lower_limit：存储大于等于下限重复次数的数据
    # path：存储到的文档路径
    # tiitle：标题数组

    # 把结果筛转换为列表存储
    rows = list()
    for key, value in result.items():
        if value >= lower_limit:
            rows.append((key, value))

    # 存入表中
    wb = openpyxl.Workbook()
    ws = wb.active
    # 填写标题
    for index in range(len(titles)):
        ws.cell(1, index+1, value=titles[index])
    # 填写内容
    for r_index in range(len(rows)):
        for c_inedx in range(len(rows[r_index])):
            ws.cell(r_index+2, c_inedx+1, value=rows[r_index][c_inedx])
    wb.save(path)
    wb.close()


# 获取号段内容
def get_haoduan(path):
    # path：号段文档路径
    result = {}
    data = xlrd.open_workbook(path)
    table = data.sheet_by_index(0)
    rows = table.nrows
    for i in range(1, rows):
        name = str(table.cell_value(i, 0))[:3].strip()
        value = table.cell_value(i, 1).strip()
        result[name] = value
    return result


def get_txt(from_path1, from_path2):
    # 假设分成两个txt
    f = open(from_path1)
    data1 = f.read()
    f.close()
    f = open(from_path2)
    data2 = f.read()
    f.close()
    data1_result = data1.split('\n')  # 可能最后会存几行空行单独删掉
    del data1_result[-1]
    del data1_result[-1]
    data2_result = data2.split('\n')
    result1 = {}
    result2 = {}
    result3 = {}
    result4 = {}
    # 1和2
    for value in data1_result:
        if value.split(',')[0] in result1:
            result1[value.split(',')[0]] += 1
        else:
            result1[value.split(',')[0]] = 1
        if value.split(',')[1] in result2:
            result2[value.split(',')[1]] += 1
        else:
            result2[value.split(',')[1]] = 1
    for value in data2_result:
        if value.split(',')[0] in result1:
            result1[value.split(',')[0]] += 1
        else:
            result1[value.split(',')[0]] = 1
        if value.split(',')[1] in result2:
            result2[value.split(',')[1]] += 1
        else:
            result2[value.split(',')[1]] = 1
    # 3
    for value in data1_result:
        if value.split(',')[0][:7] in result3:
            result3[value.split(',')[0][:7]] += 1
        else:
            result3[value.split(',')[0][:7]] = 1
    for value in data2_result:
        if value.split(',')[0][:7] in result3:
            result3[value.split(',')[0][:7]] += 1
        else:
            result3[value.split(',')[0][:7]] = 1
    # 4
    for value in data1_result:
        if value.split(',')[1][:7] in result4:
            result4[value.split(',')[1][:7]] += 1
        else:
            result4[value.split(',')[1][:7]] = 1
    for value in data2_result:
        if value.split(',')[1][:7] in result4:
            result4[value.split(',')[1][:7]] += 1
        else:
            result4[value.split(',')[1][:7]] = 1
    set_result(result1, 1, r'C:\Users\95111\Desktop\luoyuan\4_7\7\7_1.xls', ['分享手机号', '次数'])
    set_result(result2, 2, r'C:\Users\95111\Desktop\luoyuan\4_7\7\7_2.xls', ['被分享手机号', '次数'])
    set_result(result3, 1, r'C:\Users\95111\Desktop\luoyuan\4_7\7\7_3.xls', ['分享手机号段', '次数'])
    set_result(result4, 1, r'C:\Users\95111\Desktop\luoyuan\4_7\7\7_4.xls', ['被分享手机号段', '次数'])


def get1_4(lower_limit, to_path, titles, from_path, header, column_index, phone_range=[False, 0]):
    # 分析获取数据，并填写到表里
    result = get_repeat_tiame(from_path, header, column_index, phone_range)
    set_result(result, lower_limit, to_path, titles)


def get5(from_path, to_path, header, titles):
    # 统计 分享号码+时间 重复次数3次及以上
    result = get_repeat_tiame2(from_path, header)
    rows = list()
    for key, value in result.items():
        if value >= 3:
            rows.append((key.split('_')[0], key.split('_')[1], value))
    # 存入表中
    wb = openpyxl.Workbook()
    ws = wb.active
    # 填写标题
    for index in range(len(titles)):
        ws.cell(1, index + 1, value=titles[index])
    # 填写内容
    for r_index in range(len(rows)):
        for c_inedx in range(len(rows[r_index])):
            ws.cell(r_index + 2, c_inedx + 1, value=rows[r_index][c_inedx])
    wb.save(to_path)
    wb.close()


def haotou_classify(haotou_path, from_path, header, column_index, phone_range=[False, 0]):
    # 运营商分类
    from_result = get_repeat_tiame(from_path, header, column_index, phone_range)
    haotou_catalog = get_haoduan(haotou_path)
    # 统计号段重复次数
    result = {}
    for from_result_key in from_result.keys():
        try:
            if from_result_key not in haotou_catalog.keys():
                print(from_result_key+'没有')
                continue
            if haotou_catalog[from_result_key] in result.keys():
                result[haotou_catalog[from_result_key]] += from_result[from_result_key]
            else:
                result[haotou_catalog[from_result_key]] = from_result[from_result_key]
        except Exception as e:
            traceback.print_exc()
            print(e)
    return result


def all(to_path, from_paths, header, column_index, phone_range=[False, 0]):
    # 统计几个月的号码重复次数
    result4 = get_repeat_tiame(from_paths[0], header, column_index, phone_range)
    result5 = get_repeat_tiame(from_paths[1], header, column_index, phone_range)
    result6 = get_repeat_tiame(from_paths[2], header, column_index, phone_range)
    result7 = get_repeat_tiame(from_paths[3], header, column_index, phone_range)
    result8 = get_repeat_tiame(from_paths[4], header, column_index, phone_range)
    result = {}
    for key4, value4 in result4.items():
        if key4 in result.keys():
            result[key4]['4'] = value4
        else:
            result[key4] = {'4': value4}
    for key5, value5 in result5.items():
        if key5 in result.keys():
            result[key5]['5'] = value5
        else:
            result[key5] = {'5': value5}
    for key6, value6 in result6.items():
        if key6 in result.keys():
            result[key6]['6'] = value6
        else:
            result[key6] = {'6': value6}
    for key7, value7 in result7.items():
        if key7 in result.keys():
            result[key7]['7'] = value7
        else:
            result[key7] = {'7': value7}
    for key8, value8 in result8.items():
        if key8 in result.keys():
            result[key8]['8'] = value8
        else:
            result[key8] = {'8': value8}
    path = to_path
    rows = list()
    for key, value in result.items():
        if len(value) != 1:
            rows.append((key, value))
    wb = openpyxl.Workbook()
    ws = wb.active
    titles = ['被分享号码', '四月', '五月', '六月', '七月', '八月']
    for index in range(len(titles)):
        ws.cell(1, index + 1, value=titles[index])
    for r_index in range(len(rows)):
        for c_inedx in range(len(rows[r_index])):
            if c_inedx == 0:
                ws.cell(r_index + 2, c_inedx + 1, value=rows[r_index][c_inedx])
            else:
                for key_index, value_index in rows[r_index][c_inedx].items():
                    if key_index == '4':
                        ws.cell(r_index + 2, 2, value=value_index)
                    elif key_index == '5':
                        ws.cell(r_index + 2, 3, value=value_index)
                    elif key_index == '6':
                        ws.cell(r_index + 2, 4, value=value_index)
                    elif key_index == '7':
                        ws.cell(r_index + 2, 5, value=value_index)
                    elif key_index == '8':
                        ws.cell(r_index + 2, 6, value=value_index)
    wb.save(path)
    wb.close()


def share_shared_classfy(haotou_path, from_path, header):
    # 先根据分享号码分运营商，再把分好的分享号里再根据被分享号码分运营商
    haotou_catalog = get_haoduan(haotou_path)
    result = {'电信': {'总数': 0, '电信': 0, '移动': 0, '联通': 0, '虚拟': 0},
              '移动': {'总数': 0, '电信': 0, '移动': 0, '联通': 0, '虚拟': 0},
              '联通': {'总数': 0, '电信': 0, '移动': 0, '联通': 0, '虚拟': 0},
              '虚拟': {'总数': 0, '电信': 0, '移动': 0, '联通': 0, '虚拟': 0}}

    # 针对七月份txt文档数据做分析
    # f = open(r'C:\Users\95111\Desktop\luoyuan\4_7\7_1.TXT')
    # data1 = f.read()
    # f.close()
    # f = open(r'C:\Users\95111\Desktop\luoyuan\4_7\7_2.TXT')
    # data2 = f.read()
    # f.close()
    # data1_result = data1.split('\n')  # 可能最后会存几行空行单独删掉
    # del data1_result[-1]
    # del data1_result[-1]
    # data2_result = data2.split('\n')
    #
    # for value in data1_result:
    #     if value.split(',')[0][:3] not in haotou_catalog.keys():
    #         print(value.split(',')[0] + '，该分享号码没有运营商')
    #     else:
    #         result[haotou_catalog[value.split(',')[0][:3]]]['总数'] += 1
    #         if value.split(',')[1][:3] not in haotou_catalog.keys():
    #             print(value.split(',')[1] + '，该被分享号码没有运营商')
    #         else:
    #             result[haotou_catalog[value.split(',')[0][:3]]][haotou_catalog[value.split(',')[1][:3]]] += 1
    # for value in data2_result:
    #     if value.split(',')[0][:3] not in haotou_catalog.keys():
    #         print(value.split(',')[0] + '，该分享号码没有运营商')
    #     else:
    #         result[haotou_catalog[value.split(',')[0][:3]]]['总数'] += 1
    #         if value.split(',')[1][:3] not in haotou_catalog.keys():
    #             print(value.split(',')[1] + '，该被分享号码没有运营商')
    #         else:
    #             result[haotou_catalog[value.split(',')[0][:3]]][haotou_catalog[value.split(',')[1][:3]]] += 1

    # from_path：csv文档存储路径
    # header：csv文档的标题
    # column_index：统计哪一列
    # phone_range：截取号码字段
    with open(from_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if header in row[0]:
                # 是标题的话不存入结果resulr
                continue
            if row[0][:3] not in haotou_catalog.keys():
                print(row[0] + '，该分享号码没有运营商')
            else:
                result[haotou_catalog[row[0][:3]]]['总数'] += 1
                if row[1][:3] not in haotou_catalog.keys():
                    print(row[1] + '，该被分享号码没有运营商')
                else:
                    result[haotou_catalog[row[0][:3]]][haotou_catalog[row[1][:3]]] += 1
    return result


if __name__ == '__main__':
    haotou_path = r'C:\Users\95111\Desktop\luoyuan\haotou.xlsx'
    from_path = r'C:\Users\95111\Desktop\luoyuan\8.csv'
    result = share_shared_classfy(haotou_path, from_path, '手机号')
